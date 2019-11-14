from datetime import *

import openpyxl
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog
import pyexcel as p
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
import copy


app = QApplication(sys.argv)
# определяем стили
FONT = Font(name='FreeMono', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
            color='FF000000')
FILL = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='c2c2c2')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False,
                         indent=0)

BORDER = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                diagonal=Side(border_style='thin', color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='thin', color='FF000000'),
                vertical=Side(border_style='thin', color='FF000000'),
                horizontal=Side(border_style='thin', color='FF000000'))


class App(QWidget):

    def __init__(self):
        super().__init__()
        self.title = 'Эксперт-ТК / Журналы'
        self.left = 10
        self.top = 10
        self.width = 700
        self.height = 500
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        self.openFileNameDialog()

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "Загрузить файл", "/",
                                                  "All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            self.parse_file(fileName)

    def saveFileDialog(self, init, work, fire):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file, _ = QFileDialog.getSaveFileName(self, "Сохранить ", " ", "All Files (*);;Text Files (*.text)",
                                              options=options)

        if file:
            init[0].save(file + init[1])
            work[0].save(file + work[1])
            fire[0].save(file + fire[1])
        sys.exit()

    def parse_file(self, file_name):
        # Делаем проверку, если у файла расширение не xlsx, то конвертируем его в test.xlsx
        # иначе оставляем как есть
        if '.xlsx' not in file_name:
            p.save_book_as(file_name=file_name,
                           dest_file_name='test.xlsx')
            file_name = 'test.xlsx'

        # читаем excel-файл
        wb = openpyxl.load_workbook(file_name)

        # получаем активный лист
        sheet = wb.active

        # Сколько строк в файле
        rows = sheet.max_row
        # Сколько колонок в файле
        cols = sheet.max_column

        ''' 
        Создаем массив в котором хранятся ключи словаря в дальнейшем kwarg (для записи в json формат)
        date_in - дата приема на работу
        fio_staff - ФИО сотрудника
        year_birth - год рождения
        profession_staff - профессия
        department_name - наименование подразделения
        fio_instructor - ФИО инструктируемого
        leave - увольнение (если есть то это дата)
        workday - кол-во смен
        instruction_period - период инструктажа
        date_work_permission - допуск к работе
        '''

        title_list_all = ['date_in', 'fio_staff', 'year_birth', 'profession_staff', 'department_name', 'fio_instructor',
                          'leave', 'workday', 'instruction_period', 'date_work_permission']

        # Инициализируем пустой массив в котором будут хранится ассоциативные массивыт(словари)
        start_list = []
        # Инициализируем ассоциативный массив в котором будут хранится данные, а ключами для них будут служить значения
        # из массива title_list
        kwarg = dict()

        # Проходим циклом по строкам
        # Цикл начинается со второго элемента, т. к. первые 2 строчки являются заголовками
        for i in range(2, rows + 1):
            # Цикл по столбцам строки
            # Идет запись в словарь
            for j in range(1, cols + 1):
                cell = sheet.cell(row=i, column=j)
                kwarg[title_list_all[j - 1]] = cell.value
            # Вычисляем дату приема на работу
            kwarg[title_list_all[-1]] = date_calc(kwarg['date_in'], kwarg['workday']).date().strftime("%d.%m.%Y")
            # Дбавляем словарь с данными в массив и очищаем словарь
            start_list.append(kwarg)
            kwarg = {}
        init = initiative(start_list)
        work = workplace(start_list)
        fire_dict = fire(start_list)
        init_path = create_file(init['list'], init['header'], init['title'], init['col'])
        work_path = create_file(work['list'], work['header'], work['title'], work['col'])
        fire_path = create_file(fire_dict['list'], fire_dict['header'], fire_dict['title'], fire_dict['col'])
        if init_path and work_path and fire_path:
            self.saveFileDialog(init_path, work_path, fire_path)


def create_file(rows, header, title, col_list):
    # объект
    wb = Workbook()

    # активный лист
    ws = wb.active

    # название страницы
    ws.title = title

    # циклом записываем данные
    ws.append(header[0])
    ws.append(header[1])

    for row in rows:
        if type(row[0]) is not int:
            row[0] = row[0].strftime('%d.%m.%Y')
        else:
            row[1] = row[1].strftime('%d.%m.%Y')
        ws.append(row)

    # объединяем колонки
    for i in col_list:
        ws.merge_cells(i)

    # шрифт
    r = 'A1:' + col_list[-1][-2] + str(len(rows) + 2)
    for cellObj in ws[r]:
        for cell in cellObj:
            ws[cell.coordinate].font = FONT
            ws[cell.coordinate].border = BORDER
            ws[cell.coordinate].alignment = ALIGN_CENTER
            ws.column_dimensions[cell.coordinate[0]].width = len(str(cell.value)) + 15

    for i in range(1, ws.max_row + 1):
        rd = ws.row_dimensions[i]
        rd.height = 40

    # раскрашивание фона для заголовков
    for cellObj in ws['A1:' + col_list[-1][-2] + '2']:
        for cell in cellObj:
            ws[cell.coordinate].fill = FILL

    file_name = (title.replace(' ', '_') + "_" +
                 str(datetime.now().strftime("%d.%m.%Y %H:%M")) + ".xlsx")
    return [wb, file_name]


# Функции которые формируют все необходимые данные для журналов
# данные для записи в таблицу
# название журнала
# список колонок, которые нужно объеденить
# название колонок
def initiative(a):
    keys = ['date_in', 'fio_staff', 'year_birth', 'profession_staff', 'department_name', 'fio_instructor']
    res = dict()
    res_list = []
    for i in a:
        f = []
        for key in keys:
            if key != 'date_in':
                f.append(i[key])
            else:
                f.append(i[key].date())
        res_list.append(f)
    res['list'] = sorted(res_list, key=lambda d: d[0])
    res['title'] = 'Вводный инструктаж'
    res['col'] = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:H1']
    res['header'] = [('Дата', 'ФИО инструктируемого', 'Год рождения', 'Профессия, должность инструктируемого',
                      'Наименование подразделения, в которое направляется инструктируемый',
                      'Фамилия, инициалы  инструктирующего', 'Подпись'),
                     ('', '', '', '', '', '', 'Инструктирующего', 'Инструктируемого')]
    return res


def workplace(a):
    keys = ['fio_staff', 'year_birth', 'profession_staff', 'instruction_type', "", 'fio_instructor', "", "", 'workday',
            "", 'date_work_permission']
    res = dict()
    res_list = []
    for i in a:
        instruction_type = 'Первичный'
        date_in = i['date_in']
        f = [date_in.date()]
        for key in keys:
            if key != 'instruction_type':
                if key != "":
                    f.append(i[key])
                else:
                    f.append("")
            else:
                f.append(instruction_type)

        res_list.append(f)
        while date_in.date() < datetime.today().date():
            f1 = copy.deepcopy(f)
            if i['leave'] is not None and i['leave'].date() < date_in.date():
                pass
            instruction_type = "Повторный"
            date_in += timedelta(days=int(i['instruction_period']) * 365 / 12)
            while parse_holiday(date_in):
                date_in -= timedelta(days=1)
            if date_in.date() < datetime.today().date():
                f1[0] = date_in.date()
                f1[4] = instruction_type
                f1[9] = ""
                f1[-1] = ""
                res_list.append(f1)
    res['list'] = sorted(res_list, key=lambda d: d[0])
    res['title'] = 'Инструктаж на рабочем месте'
    res['col'] = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:I1', 'J1:L1']
    res['header'] = [('Дата', 'ФИО инструктируемого', 'Год рождения', 'Профессия, должность инструктируемого',
                      'Вид инструктажа', 'Причина проведения внепланового инструктажа',
                      'Фамилия, инициалы  инструктирующего', 'Подпись', '', 'Стажировка на рабочем месте', ''),
                     ('', '', '', '', '', '', '', 'Инструктирующего', 'Инструктируемого', 'Количество смен',
                      'Стажировку прошел (подпись работника)', 'Допуск к работе')]
    return res


def fire(a):
    keys = ['fio_staff', 'year_birth', 'profession_staff', 'instruction_type', "department_name", 'fio_instructor', "",
            ""]
    res = dict()
    res_list = []
    for i in a:
        instruction_type = 'Вводный'
        c = 0
        date_in = i['date_in']
        while (c < 2):
            f = [0]
            f.append(date_in.date())
            for key in keys:
                if key != 'instruction_type':
                    if key != "":
                        f.append(i[key])
                    else:
                        f.append("")
                else:
                    if c == 1:
                        instruction_type = 'Первичный'
                    f.append(instruction_type)
            res_list.append(f)
            c += 1
        while date_in.date() < datetime.today().date():
            f1 = copy.deepcopy(f)
            if i['leave'] is not None and i['leave'].date() < date_in.date():
                pass
            instruction_type = "Повторный"
            date_in += timedelta(days=int(i['instruction_period']) * 365 / 12)
            while parse_holiday(date_in):
                date_in -= timedelta(days=1)
            if date_in.date() < datetime.today().date():
                f1[1] = date_in.date()
                f1[5] = instruction_type
                f1[8] = ""
                f1[-1] = ""
                res_list.append(f1)
    res_list = sorted(res_list, key=lambda d: d[1])
    for v in range(len(res_list)):
        res_list[v][0] = v + 1
    res['list'] = res_list
    res['title'] = 'Инструктаж по пожарной безопасности'
    res['col'] = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:H2', 'I1:J1']
    res['header'] = [('№ п\п', 'Дата', 'ФИО инструктируемого', 'Год рождения', 'Профессия, должность инструктируемого',
                      'Вид инструктажа', 'Подразделение, куда направляется инструктируемый',
                      'Фамилия, инициалы  инструктирующего', 'Подпись'),
                     ('', '', '', '', '', '', '', '', 'Инструктирующего', 'Инструктируемого')]
    return res


def date_calc(day, wd):
    while parse_holiday(day):
        day += timedelta(days=1)  # прибавляем 1 день до тех пор пока не станет будним д.н.
    for i in range(1, wd):
        day += timedelta(days=1)
        while parse_holiday(day):
            day += timedelta(days=1)  # прибавляем 1 день до тех пор пока не станет будним д.н.
    return day


def parse_holiday(next_date):
    year = next_date.year
    month = next_date.month
    day = next_date.day
    wb = openpyxl.load_workbook(filename="holiday.xlsx")
    sheet = wb.active
    for row in sheet.rows:
        if row[0].value == year:
            if str(day) in row[month].value.split():
                return True
    return False


ex = App()
sys.exit(app.exec_())
